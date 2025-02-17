import os
import sys
import shutil
import pandas as pd
from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
import xlsxwriter

# 🟢 Poprawione: Pobranie poprawnej ścieżki do katalogu aplikacji (.exe i dev)
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)  # Jeśli uruchamiamy jako .exe, to pobieramy folder, w którym jest plik exe
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # Standardowa ścieżka w trybie deweloperskim

TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")  
STATIC_DIR = os.path.join(BASE_DIR, "static")  

# 🟢 Poprawione: Poprawne ścieżki do plików Excel
FILE_PATH = os.path.join(BASE_DIR, "export_order_articles.xlsx")  
DUPLICATE_FILE_PATH = os.path.join(BASE_DIR, "export_order_articles_filtered.xlsx")

# 🟢 Sprawdzenie, czy plik Excel istnieje
if not os.path.exists(FILE_PATH):
    raise FileNotFoundError(f"❌ BŁĄD: Nie znaleziono pliku {FILE_PATH}. Upewnij się, że plik jest w tym samym katalogu co `app.exe`!")

# 🟢 Inicjalizacja aplikacji Flask
app = Flask(__name__, template_folder=TEMPLATES_DIR, static_folder=STATIC_DIR)

# Wczytanie danych z Excela
df = pd.read_excel(FILE_PATH, sheet_name="Worksheet", dtype=str)
df.columns = df.columns.str.strip()  # Normalizacja kolumn

@app.route("/")
def index():
    return render_template("form.html")

@app.route("/autocomplete", methods=["GET"])
def autocomplete():
    query = request.args.get("query", "").strip().lower()
    if not query:
        return jsonify([])

    suggestions = df["Article-Nr."].astype(str).str.strip()
    filtered = suggestions[suggestions.str.lower().str.contains(query)].unique().tolist()
    
    return jsonify(filtered)

@app.route("/get_variants", methods=["GET"])
def get_variants():
    article = request.args.get("article", "").strip()
    if not article:
        return jsonify({"colors": [], "sizes": []})

    filtered_df = df[df["Article-Nr."].astype(str).str.strip() == article]

    available_colors = filtered_df["Color"].astype(str).str.strip().unique().tolist()
    available_sizes = filtered_df["Size"].astype(str).str.strip().unique().tolist()

    return jsonify({"colors": available_colors, "sizes": available_sizes})

@app.route("/update_quantity", methods=["POST"])
def update_quantity():
    try:
        data = request.json
        article = data.get("article", "").strip()
        color = data.get("color", "").strip()
        size = str(data.get("size", "")).strip()
        quantity = int(data.get("quantity", 0)) if str(data.get("quantity", "0")).isdigit() else 0

        if not article or not color or not size or quantity <= 0:
            return jsonify({"status": "error", "message": "Brak wymaganych danych lub ilość musi być większa od 0!"}), 400

        mask = (df["Article-Nr."].astype(str).str.strip() == article) & \
               (df["Color"].astype(str).str.strip() == color) & \
               (df["Size"].astype(str).str.strip() == size)

        if mask.any():
            df.loc[mask, "Quantity"] = df.loc[mask, "Quantity"].astype(int) + quantity
            df.to_excel(FILE_PATH, index=False, sheet_name="Worksheet", engine="openpyxl")
            return jsonify({"status": "success", "message": "Ilość została zaktualizowana!"})
        else:
            return jsonify({"status": "error", "message": "Nie znaleziono artykułu!"}), 404

    except Exception as e:
        return jsonify({"status": "error", "message": f"Błąd serwera: {str(e)}"}), 500

@app.route("/duplicate_and_remove_zero", methods=["POST"])
def duplicate_and_remove_zero():
    try:
        # Kopiowanie pliku do nowej wersji
        shutil.copy(FILE_PATH, DUPLICATE_FILE_PATH)
        df_copy = pd.read_excel(DUPLICATE_FILE_PATH, sheet_name="Worksheet", dtype=str)
        df_copy.columns = df_copy.columns.str.strip()
        df_copy = df_copy[df_copy["Quantity"].astype(int) > 0]

        # Ustawienie wszystkich kolumn jako string
        for col in df_copy.columns:
            df_copy[col] = df_copy[col].astype(str)

        # Usunięcie ukrytych metadanych i pełna konwersja pliku
        with pd.ExcelWriter(DUPLICATE_FILE_PATH, engine="xlsxwriter") as writer:
            df_copy.to_excel(writer, index=False, sheet_name="Worksheet")

        # Otwórz plik w `openpyxl`, aby poprawnie zapisać format `.xlsx`
        wb = load_workbook(DUPLICATE_FILE_PATH)
        ws = wb["Worksheet"]

        # Formatowanie wszystkich komórek jako "Zwykły tekst"
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = "@"

        # Poprawny zapis, aby plik był zgodny z Excelem
        wb.save(DUPLICATE_FILE_PATH)
        wb.close()

        # Zapisz jeszcze raz w pełnym formacie `.xlsx`
        df_clean = pd.read_excel(DUPLICATE_FILE_PATH, sheet_name="Worksheet", dtype=str)
        with pd.ExcelWriter(DUPLICATE_FILE_PATH, engine="openpyxl") as writer:
            df_clean.to_excel(writer, index=False, sheet_name="Worksheet")

        # Wyzerowanie ilości w oryginalnym pliku
        global df
        df["Quantity"] = "0"
        with pd.ExcelWriter(FILE_PATH, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Worksheet")

        return jsonify({"status": "success", "message": "Plik został poprawnie przekonwertowany na w pełni działający `.xlsx`!"})

    except Exception as e:
        return jsonify({"status": "error", "message": f"Błąd serwera: {str(e)}"}), 500

if __name__ == "__main__":
    app.run(debug=True, port=5001)
